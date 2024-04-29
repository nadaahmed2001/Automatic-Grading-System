# OCR_Model.py

import requests
import json
from openpyxl import load_workbook

def ocr_space_file(filename, overlay=False, api_key='1b70baf52f88957', language='eng'):
    payload = {'isOverlayRequired': overlay,
               'apikey': api_key,
               'language': language,
               'detectOrientation': 'true',
               'scale' : 'true',
               'OCREngine': 2,
              }
    with open(filename, 'rb') as f:
        r = requests.post('https://api.ocr.space/parse/image',
                          files={filename: f},
                          data=payload,
                          )
    return r.content.decode()

def extract_text_and_store(images_excel):
    workbook = load_workbook(images_excel)
    sheet = workbook['StudentAnswerExtracted']

    if sheet.cell(row=1, column=2).value is None:
        sheet.cell(row=1, column=2, value="Extracted Text")

    for row in range(2, sheet.max_row + 1):
        image_path = sheet.cell(row=row, column=1).value
        extracted_text = ocr_space_file(filename=image_path)

        response_data = json.loads(extracted_text)
        extracted_text_concatenated = response_data["ParsedResults"][0]["ParsedText"] if "ParsedResults" in response_data else ""

        lines = extracted_text_concatenated.split('\n')
        
        first_line = lines[0] if lines else ""
        extracted_text_concatenated = '\n'.join(lines[1:]) if lines else ""

        sheet.cell(row=row, column=2, value=extracted_text_concatenated)
        sheet.cell(row=row, column=3, value=first_line)

    workbook.save(images_excel)
    return f"Extracted text stored in '{images_excel}'."

if __name__ == "__main__":
    file_path = 'Final_SBERT_Dataset.xlsx'
    print(extract_text_and_store(file_path))
